# -*- coding: utf-8 -*

# [03.엑셀로 카카오플러스 기본 UI 구현하기 - 엑셀챗봇빌더]
# 파이썬 기본 문법으로 구현된 카카오플러스 기본 UI 기능을
# 엑셀에 데이터 입력을 통해 구현되도록 해본다.
# 즉, 엑셀을 통한 챗봇 빌더 기능을 구현해본다.
from flask import Flask, request, jsonify, json
from openpyxl import load_workbook, cell
import excel_db

app = Flask(__name__)

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)

user_db = db['User']


@app.route("/keyboard")
def Keyboard():

    response = {
        "message" : {
            "text" : "원하시는 버튼을 눌러주세요."
        },
        "type" : "buttons",
        "buttons" : ["홈으로"]
    }
    return jsonify(response)


@app.route("/message", methods=["POST"])
def message():
    data = json.loads(request.data)
    content = data["content"]
    user_key = data["user_key"]

    # [02.엑셀로 사용자 정보 관리하기]
    for idx, row in enumerate(user_db.rows):
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break

        if idx == user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]

            response = {
                "message" : {
                    "text" : "처음 오셨네요? 이름이 뭐에요?"
                },
                "keyboard" : {
                    "type": "text"
                }
            }
            return jsonify(response)

    if user_row[1].value is 0:
        user_row[1].value = 1
        user_row[2].value = content
        db.save(EXCEL_FILE_NAME)

        response = {
            "message" : {
                "text" : "{name}님, 반갑습니다.".format(name=content)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

        return jsonify(response)

    try:
        # [03.엑셀로 카카오플러스 기본 UI 구현하기 - 엑셀챗봇빌더]
        response = excel_db.get_response(content, user_row)

    # 예기치 않은 에러발생시, 아래와 같은 메시지를 사용자에게 전달한다.
    except:
        response = {
            "message" : {
                "text" : "다시 시도해 주세요."
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

    return jsonify(response)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80, debug=True)
