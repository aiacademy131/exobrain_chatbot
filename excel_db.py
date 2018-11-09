from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict

db = load_workbook(filename='Database.xlsx')
seminar_db = db['Seminar']
lecture_db = db['Lecture']
user_db = db['User_data']
answer_db = db['Answer']

### 셀 병합시 첫 행을 제외한 나머지는 None으로 출력
### 병합셀에 대한 처리 필요

#설명회 주제 출력
# 분류 라인이 일정하다고 가정



def seminar_topic():
    contents = []

    #설명 라인 생략
    for row in range(5, seminar_db.max_row):

        content = seminar_db[row][0].value
        contents.append(content)

    return contents



def seminar_info(text):

    info_sets = []

    for i in seminar_db.iter_rows(min_row=5, max_col=5):
        if text == i[0].value:
            for k in seminar_db.iter_cols(min_row=4):
                type = k[0].value + " : "
                content = k[2].value + "\n"
                info_set = (type + content)
                info_sets.append(info_set)

    return "".join(info_sets)

#큰 주제를 출력
def lecture_topic():
    contents = []

    for row in lecture_db.iter_rows(min_row=1, max_col=1):
        if '[' in row[0].value :
            content = row[0].value
            contents.append(content)

    return contents

#큰 주제 내부 강의 출력
def lecture_info(topic=None):

    info_dic = {}
    info_sets = []
    next_lecture_row = lecture_db.max_row + 1
    lectureRow = 3

    if topic is None:
        for row in lecture_db.iter_rows(min_row=lectureRow, max_col=1):
            info_sets.append(row[0].value)
        return info_sets

    elif topic == "bayesian":
        for row in lecture_db:
            if '[' in row[0].value:
                info_dic = (row[0].value,row[1].value.split('.'))
                info_sets.append(info_dic)
        return info_sets

    for row in lecture_db:
        #print(row[0].value)
        if row[0].value == topic:

            lectureRow = row[0].row

    for row in lecture_db.iter_rows(min_row=lectureRow + 1, max_col=1):
        if '[' in row[0].value:
            next_lecture_row = row[0].row
            break

    for row in lecture_db.iter_rows(min_row=lectureRow + 1, max_row =next_lecture_row - 1):
        info_sets.append(row[0].value)

    return info_sets

#강의 정보 표시
def detail_lecture_info(content):
    info_sets = []
    for row in lecture_db:
        if row[0].value == content:
            for k in range(lecture_db.max_column - 1):     #마지막 난이도는 제외
                type = lecture_db[2][k].value + " : "
                info = row[k].value + "\n"
                info_set = (type + info)
                info_sets.append(info_set)
    return "".join(info_sets)
#답변한 데이터 저장

def save_answer_data(Answer, depth, user_row):

    user_db[user_row][depth - 2].value = Answer
    db.save(filename='Database.xlsx')

#엑셀에 있는 질문 뽑아오기
def question_msg(user_type, depth):
    for row in answer_db:
        if row[0].value == depth and user_type in row[1].value.split(" "):
            return row[3].value


def return_user_type(user_row):

    return user_db[user_row][1].value


#질문에 따른 사용자 정보 불러오기
def custom_lecture_info(user_row):

    user_level = 0
    custom_lectures = []

    for col in range(1, 4):
        if user_db[2][col].value not in ["고등학생","아니오", None]:
            user_level += 1

    for row in lecture_db:
        if row[7].value == user_level:
            custom_lectures.append(row[0].value)

    return custom_lectures

#QuestionDepth 증가
def next_question_depth(user_row):

    user_db[user_row][6].value += 1
    db.save(filename='Database.xlsx')


#현재 상태의 QuestionDepth 리턴
def return_question_depth(user_row):

    return user_db[user_row][6].value

#QuestionDepth 초기화
def initalize_question_depth(user_row, depth):

    user_db[user_row][6].value = depth
    db.save(filename='Database.xlsx')


#Excel sheet에서 content와 일치하는 row.value 리턴
def excel_ui(content):
    answer_row_data = []
    excel_question_list = []
    for row in range(answer_db.max_row - 1):
        if answer_db[row + 2][1].value is not None:
            excel_question_content = answer_db[row + 2][1].value.split(" - ")
            excel_question_list.extend(excel_question_content)

    excel_question_list = list(OrderedDict.fromkeys(excel_question_list))
    for idx in range(len(excel_question_list)):
        if excel_question_list[idx] == content:
            for i in range(answer_db.max_column):
                if answer_db[idx + 2][i].value is not None:
                    answer_row_data.append(answer_db[idx + 2][i].value)

    return answer_row_data


#사용자 정보 확인 후 사용자 row return
def find_user_row(userkey):
    user_exist = False

    for row in user_db.rows:  # --> findUserkey()
        if row[0].value == userkey:
            user_exist = True
            user_row = row[0].row       # tuple


    if user_exist == False:         # --> createUserInfo()
        user_db[user_db.max_row + 1][0].value = userkey
        # print(user_db[user_db.max_row][0].value)
        # 중간에 데이터 값을 지울때는 행 자체를 삭제해야 빈칸이 안생긴다.
        user_db[user_db.max_row][6].value = 0  # Question Depth Initialize
        user_row = user_db.max_row
        db.save(filename='Database.xlsx')

    return user_row

#t사용자 데이터 삭제
def delete_userdata(user_row):
    user_db.delete_rows(user_row, 1)
    db.save(filename='Database.xlsx')

def contents_list(question_depth):
    contents = []

    for row in range(answer_db.max_row):
        if answer_db[row + 2][0].value == question_depth:
            content = answer_db[row + 2][1].value
            contents.append(content)

    return contents

def save_user_topic(user_row, topic):

    for col in range(user_db.max_column):
        if "Topic" == user_db[1][col].value:
            user_db[user_row][col].value = topic
            db.save(filename='Database.xlsx')

def question_info(question_depth):

    questions = []

    for row in range(answer_db.max_row):
        if answer_db[row + 1][0].value == question_depth:
            question = answer_db[row + 1][1].value.split(" - ")
            questions.extend(question)

    for idx in range(len(contents_list(question_depth-1))):
        question_list = list(set(questions))
        try:
            question_list.remove(contents_list(question_depth-1)[idx])
        except:
            continue

    return question_list

