# -*- coding: utf-8 -*

# [05.엑소브레인 형태소 분석 API 와 머신러닝 적용하여 강좌 추천하기]
from flask import Flask, request, jsonify, json
from openpyxl import load_workbook

# 나이브 베이즈 분류 (Naive Bayesian classification)
# 데이터셋의 모든 특징들이 동등하고 독립적이라고 가정
# 예를들어 비가 오는 날에는 시간보다는 습도가 더 중요한 변수가 될 수 있지만
# 나이브베이지안 에서는 이런 사실을 무시
# 이런 가정에도 불구하고 분류학습에서 매우 정확한 결과를 내놓음
from bayesian import Filter

import excel_db

app = Flask(__name__)

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)

user_db = db['User']

USER_STATE_FIRST_JOIN = 0
USER_STATE_DEFAULT = 1
USER_STATE_REQUEST_LECTURE = 2


@app.route("/keyboard")
def Keyboard():

    response = {
        "message" : {
            "text" : "원하시는 버튼을 눌러주세요."
        },
        "type" : "buttons",
        "buttons" : ["홈으로"]
    }
    return jsonify(response)


@app.route("/message", methods=["POST"])
def message():
    data = json.loads(request.data)
    content = data["content"]
    user_key = data["user_key"]

    # [02.엑셀로 사용자 정보 관리하기]
    for idx, row in enumerate(user_db.rows):
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break

        if idx == user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]

            response = {
                "message" : {
                    "text" : "처음 오셨네요? 이름이 뭐에요?"
                },
                "keyboard" : {
                    "type": "text"
                }
            }
            return jsonify(response)

    if user_row[1].value is USER_STATE_FIRST_JOIN:
        response = {
            "message" : {
                "text" : "{name}님, 반갑습니다.".format(name=content)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

        user_row[1].value = USER_STATE_DEFAULT
        user_row[2].value = content
        db.save(EXCEL_FILE_NAME)
        return jsonify(response)

    # [05.엑소브레인 형태소 분석 API 와 머신러닝 적용하여 강좌 추천하기]
    if user_row[1].value is USER_STATE_REQUEST_LECTURE:

        # bayesian Filter의 인스턴스 객체를 가져옴
        bf = Filter()

        # 엑셀에서 학습을 위한 데이터를 가져옴
        b_data = excel_db.get_text_for_ml()

        # 데이터 학습시키기
        for i in range(len(b_data)):
            for k in range(len(b_data[i][1][:-1])):
                bf.fit(b_data[i][1][k], b_data[i][0])

        # 사용자의 입력 문구에서 적합한 강좌 추천하기
        pre, scorelist = bf.predict(content)
        response = {
            "message" : {
                "text" : "{pre} 강좌 추천드립니다.".format(pre=pre)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

        user_row[1].value = USER_STATE_DEFAULT
        db.save(EXCEL_FILE_NAME)
        return jsonify(response)

    try:
        # [04.사용자에게 강좌 추천하기]
        if content == u"수업소개":
            if user_row[3].value is not None:
                level = user_row[3].value
                response = excel_db.get_lectures(level, user_row)
            else:
                response = {
                    "message" : {
                        "text" : "학습 수준을 알려주세요."
                    },
                    "keyboard" : {
                        "type" : "buttons",
                        "buttons" : ["초급", "중급", "고급"]
                    }
                }
        elif content in ["초급", "중급", "고급"]:
            user_row[3].value = content
            db.save(EXCEL_FILE_NAME)
            response = excel_db.get_lectures(content, user_row)

        # [05.엑소브레인 형태소 분석 API 와 머신러닝 적용하여 강좌 추천하기]
        elif content == u"수업소개(대화)":
            user_row[1].value = USER_STATE_REQUEST_LECTURE
            db.save(EXCEL_FILE_NAME)
            response = {
                "message" : {
                    "text" : "관심있는 주제를 알려주세요."
                },
                "keyboard" : {
                    "type" : "text"
                }
            }
        else:
            # [03.엑셀로 카카오플러스 기본 UI 구현하기 - 엑셀챗봇빌더]
            response = excel_db.get_response(content, user_row)

    except:
        response = {
            "message" : {
                "text" : "다시 시도해 주세요."
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

    return jsonify(response)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80, debug=True)
