# -*- coding: utf-8 -*

# [06.엑소브레인 개체명 인식 API 와 예약관리]
from flask import Flask, request, jsonify, json
from openpyxl import load_workbook, cell
from exobrain_api import exobrainNLU, get_date_from_sentence
from bayesian import Filter

import excel_db

app = Flask(__name__)

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)

user_db = db['User']

USER_STATE_FIRST_JOIN = 0
USER_STATE_DEFAULT = 1
USER_STATE_REQUEST_LECTURE = 2
USER_STATE_REQUEST_RESERVATION = 3


@app.route("/keyboard")
def Keyboard():

    response = {
        "message" : {
            "text" : "원하시는 버튼을 눌러주세요."
        },
        "type" : "buttons",
        "buttons" : ["홈으로"]
    }
    return jsonify(response)


@app.route("/message", methods=["POST"])
def message():
    data = json.loads(request.data)
    content = data["content"]
    user_key = data["user_key"]

    # [02.엑셀로 사용자 정보 관리하기]
    for idx, row in enumerate(user_db.rows):
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break

        if idx == user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]

            response = {
                "message" : {
                    "text" : "처음 오셨네요? 이름이 뭐에요?"
                },
                "keyboard" : {
                    "type": "text"
                }
            }
            return jsonify(response)

    if user_row[1].value is USER_STATE_FIRST_JOIN:
        response = {
            "message" : {
                "text" : "{name}님, 반갑습니다.".format(name=content)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

        user_row[1].value = USER_STATE_DEFAULT
        user_row[2].value = content
        db.save(EXCEL_FILE_NAME)
        return jsonify(response)

    # [05.엑소브레인 형태소 분석 API 와 베이지안 알고리즘 적용]
    if user_row[1].value is USER_STATE_REQUEST_LECTURE:
        bf = Filter()
        b_data = excel_db.get_text_for_ml()

        for i in range(len(b_data)):
            for k in range(len(b_data[i][1][:-1])):
                bf.fit(b_data[i][1][k], b_data[i][0])

        pre, scorelist = bf.predict(content)
        response = {
            "message" : {
                "text" : "{pre} 강좌 추천드립니다.".format(pre=pre)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

        user_row[1].value = USER_STATE_DEFAULT
        db.save(EXCEL_FILE_NAME)
        return jsonify(response)

    # [06.엑소브레인 개체명 인식 API 와 예약관리]
    if user_row[1].value is USER_STATE_REQUEST_RESERVATION:
        date_sentence = []

        # 엑소브레인 개체명 인식 API 를 사용하여, 문장에서 날짜, 시간, 장소정보 추출하기
        date_set = get_date_from_sentence(content)
        for k in range(len(date_set)):
            date_text = "".join(date_set[k][1]) + " : " + "".join(date_set[k][0]) + "\n"
            date_sentence.append(date_text)

        date_sentence = "".join(date_sentence)
        response = {
            "message" : {
                "text" : date_sentence
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

        user_row[1].value = USER_STATE_DEFAULT
        db.save(EXCEL_FILE_NAME)
        return jsonify(response)

    try:
        # [04.사용자에게 강좌 추천하기]
        if content == u"수업소개":
            if user_row[3].value is not None:
                level = user_row[3].value
                response = excel_db.get_lectures(level, user_row)
            else:
                response = {
                    "message" : {
                        "text" : "학습 수준을 알려주세요."
                    },
                    "keyboard" : {
                        "type" : "buttons",
                        "buttons" : ["초급", "중급", "고급"]
                    }
                }

        elif content in ["초급", "중급", "고급"]:
            user_row[3].value = content
            db.save(EXCEL_FILE_NAME)
            response = excel_db.get_lectures(content, user_row)

        # [05.엑소브레인 형태소 분석 API 와 베이지안 알고리즘 적용]
        elif content == u"수업소개(대화)":
            user_row[1].value = USER_STATE_REQUEST_LECTURE
            db.save(EXCEL_FILE_NAME)

            response = {
                "message" : {
                    "text" : "관심있는 주제를 알려주세요."
                },
                "keyboard" : {
                    "type" : "text"
                }
            }

        # [06.엑소브레인 개체명 인식 API 와 예약관리]
        elif content == u"사전예약":
            user_row[1].value = USER_STATE_REQUEST_RESERVATION
            db.save(EXCEL_FILE_NAME)

            response = {
                "message" : {
                    "text" : "예약 시간을 입력해 주세요."
                },
                "keyboard" : {
                    "type" : "text"
                }
            }
        else:
            # [03.엑셀로 카카오플러스 기본 UI 구현하기 - 엑셀챗봇빌더]
            response = excel_db.get_response(content, user_row)

    except:
        response = {
            "message" : {
                "text" : "다시 시도해 주세요."
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

    return jsonify(response)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80, debug=True)
