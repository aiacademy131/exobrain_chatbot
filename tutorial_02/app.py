# -*- coding: utf-8 -*

# [02.엑셀로 사용자 정보 관리하기]
# 이용자가 최초로 채팅방에 들어올 때 사용자 정보가 없는지 확인
# 처음으면 사용자의 이름을 묻고 해당 정보를 엑셀에 저장한다.
# 파이썬 엑셀파일 처리 라이브러리(openpyxl)사용법을 이해한다.
from flask import Flask, request, jsonify, json
from openpyxl import load_workbook, cell

app = Flask(__name__)

# 사용자 정보를 관리할 엑셀파일이름
EXCEL_FILE_NAME = 'Database.xlsx'

# openpyxl.load_workbook(엑셀파일명)함수 호출로 Workbook 객체를 얻는다.
db = load_workbook(filename=EXCEL_FILE_NAME)

# 엑셀파일에서 "User" 시트 접근 가능한 객체를 가져온다.
user_db = db['User']


@app.route("/keyboard")
def Keyboard():

    response = {
        "message" : {
            "text" : "원하시는 버튼을 눌러주세요."
        },
        "type" : "buttons",
        "buttons" : ["홈으로"]
    }
    return jsonify(response)


@app.route("/message", methods=["POST"])
def message():
    data = json.loads(request.data)
    content = data["content"]

    # 플러스친구 API 서버에서 사용자의 키값을 request 에 담아 보내준다.
    # 사용자 키값은 request.data.user_key 에 담겨있다.
    user_key = data["user_key"]

    # [02.엑셀로 사용자 정보 관리하기]
    for idx, row in enumerate(user_db.rows):
        # 엑셀 User 시트의 사용자정보 목록에서 현재 사용자의 키값이 저장되어 있으면
        # 기존 사용자라 판단하고 해당 객체를 user_row 변수에 보관한다.
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break

        # 엑셀 User 시트의 사용자정보 목록에서 현재 사용자의 키값이 없다면
        # 신규 사용자라 판단하고 해당 정보를 엑셀에 '저장'하고
        # 해당 객체를 user_row 변수에 보관한다.
        if idx == user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]

            response = {
                "message" : {
                    "text" : "처음 오셨네요? 이름이 뭐에요?"
                },
                "keyboard" : {
                    "type": "text"
                }
            }
            return jsonify(response)

    # user_row[1] 은 엑셀의 User 시트에서 두번째 열 '상태'를 가리킨다.
    # User 시트에서 상태값을 체크하여
    # 사용자의 이름을 엑셀 User 시트에 저장한다.
    if user_row[1].value is 0:
        user_row[1].value = 1
        user_row[2].value = content
        db.save(EXCEL_FILE_NAME)

        response = {
            "message" : {
                "text" : "{name}님, 반갑습니다.".format(name=content)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

        return jsonify(response)

    # [01.카카오플러스 기본 UI 구현하기]
    if content == u"홈으로":
        response = {
            "message" : {
                "text" : "원하시는 정보 버튼을 눌러주세요."
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

    elif content == u"학원소개":
        response = {
            "message" : {
                "text": "우리는 인공지능 기술을 교육합니다."
            },
            "keyboard" : {
                "type" : "buttons",
                "buttons" : ["홈으로"]
            }
        }

    elif content == u"시간표":
        response = {
            "message" : {
                "text" : "시간표입니다.",
                "photo" : {
                    "url" : "http://twinkleballet.co.kr/data/designImages/BANNER1_1461112781_%EC%8B%9C%EA%B0%84%ED%91%9C.jpg",
                    "width" : 640,
                    "height" : 480
                },
                "message_button" : {
                    "label" : "웹에서 시간표 보기",
                    "url" : "https://www.ai-academy.ai/ai-b2c"
                }
            },
            "keyboard" : {
                "type" : "buttons",
                "buttons" : ["홈으로"]
            }
        }

    return jsonify(response)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80, debug=True)
