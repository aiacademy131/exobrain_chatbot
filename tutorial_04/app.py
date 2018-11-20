# -*- coding: utf-8 -*

# [04.사용자에게 강좌 추천하기]
from flask import Flask, request, jsonify, json
from openpyxl import load_workbook
import excel_db

app = Flask(__name__)

EXCEL_FILE_NAME = 'Database.xlsx'
db = load_workbook(filename=EXCEL_FILE_NAME)

user_db = db['User']


@app.route("/keyboard")
def Keyboard():

    response = {
        "message" : {
            "text" : "원하시는 버튼을 눌러주세요."
        },
        "type" : "buttons",
        "buttons" : ["홈으로"]
    }
    return jsonify(response)


@app.route("/message", methods=["POST"])
def message():
    data = json.loads(request.data)
    content = data["content"]
    user_key = data["user_key"]

    # [02.엑셀로 사용자 정보 관리하기]
    for idx, row in enumerate(user_db.rows):
        if idx != 0 and row[0].value == user_key:
            user_row = row
            break

        if idx == user_db.max_row - 1:
            NEW_INDEX = user_db.max_row + 1
            user_db[NEW_INDEX][0].value = user_key
            user_db[NEW_INDEX][1].value = 0
            db.save(EXCEL_FILE_NAME)
            user_row = user_db[user_db.max_row]

            response = {
                "message" : {
                    "text" : "처음 오셨네요? 이름이 뭐에요?"
                },
                "keyboard" : {
                    "type": "text"
                }
            }
            return jsonify(response)

    if user_row[1].value is 0:
        user_row[1].value = 1
        user_row[2].value = content
        db.save(EXCEL_FILE_NAME)

        response = {
            "message" : {
                "text" : "{name}님, 반갑습니다.".format(name=content)
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["학원소개", "시간표", "홈으로"]
            }
        }

        return jsonify(response)

    try:

        # [04.사용자에게 강좌 추천하기]
        if content == u"수업소개":
            # 엑셀 User 시트에서 '수준'열의 값이 있다면,
            # 엑셀에서 사용자의 '수준'에 해당되는 Lecture 정보를 가져온다.
            if user_row[3].value is not None:
                level = user_row[3].value
                response = excel_db.get_lectures(level, user_row)

            # 엑셀 User 시트에서 '수준'열의 값이 없다면,
            # 사용자의 학습 수준을 물어본다.
            else:
                response = {
                    "message" : {
                        "text" : "학습 수준을 알려주세요."
                    },
                    "keyboard" : {
                        "type" : "buttons",
                        "buttons" : ["초급", "중급", "고급"]
                    }
                }

        # 엑셀에서 사용자의 '수준'에 해당되는 Lecture 정보를 가져온다.
        elif content in ["초급", "중급", "고급"]:
            user_row[3].value = content
            db.save(EXCEL_FILE_NAME)
            response = excel_db.get_lectures(content, user_row)

        else:
            # [03.엑셀로 카카오플러스 기본 UI 구현하기 - 엑셀챗봇빌더]
            response = excel_db.get_response(content, user_row)

    except:
        response = {
            "message" : {
                "text" : "다시 시도해 주세요."
            },
            "keyboard" : {
                "type": "buttons",
                "buttons": ["홈으로"]
            }
        }

    return jsonify(response)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80, debug=True)
